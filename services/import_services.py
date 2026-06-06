from openpyxl import load_workbook

from django.db import transaction
from django.utils.text import slugify
from django.core.exceptions import ValidationError

from .models import (
    MainService,
    Service,
    ServiceSection,
)


class ServiceExcelImporter:

    REQUIRED_SHEETS = [
        "main_services",
        "services",
        "sections",
    ]

    def __init__(self, file):
        self.file = file
        self.workbook = None

    # =====================================================
    # LOAD WORKBOOK
    # =====================================================

    def load(self):

        try:
            self.workbook = load_workbook(
                filename=self.file,
                data_only=True,
            )

        except Exception as e:
            raise ValidationError(
                f"Invalid excel file: {str(e)}"
            )

        for sheet in self.REQUIRED_SHEETS:

            if sheet not in self.workbook.sheetnames:

                raise ValidationError(
                    f"Missing required sheet: {sheet}"
                )

    # =====================================================
    # HELPERS
    # =====================================================

    def get_rows(self, sheet_name):

        sheet = self.workbook[sheet_name]

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return []

        headers = rows[0]

        data = []

        for row in rows[1:]:

            item = {}

            for idx, header in enumerate(headers):

                item[header] = row[idx]

            data.append(item)

        return data

    # =====================================================
    # MAIN SERVICES
    # =====================================================

    def import_main_services(self):

        rows = self.get_rows("main_services")

        created = 0

        for index, row in enumerate(rows, start=1):

            code = str(row.get("code", "")).strip()

            title_en = str(
                row.get("title_en", "")
            ).strip()

            title_ar = str(
                row.get("title_ar", "")
            ).strip()

            order = row.get("order") or index

            if not code:
                raise ValidationError(
                    "Main service code is required"
                )

            if not title_en:
                raise ValidationError(
                    f"Missing title_en for code {code}"
                )

            slug = slugify(title_en)

            obj, is_created = MainService.objects.update_or_create(
                code=code,
                defaults={
                    "title_en": title_en,
                    "title_ar": title_ar,
                    "slug": slug,
                    "order": order,
                    "is_active": True,
                }
            )

            if is_created:
                created += 1

        return created

    # =====================================================
    # SERVICES
    # =====================================================

    def import_services(self):

        rows = self.get_rows("services")

        created = 0

        for index, row in enumerate(rows, start=1):

            main_code = str(
                row.get("main_code", "")
            ).strip()

            title_en = str(
                row.get("title_en", "")
            ).strip()

            title_ar = str(
                row.get("title_ar", "")
            ).strip()

            short_description_en = str(
                row.get("short_description_en", "")
            ).strip()

            short_description_ar = str(
                row.get("short_description_ar", "")
            ).strip()

            order = row.get("order") or index

            if not main_code:
                raise ValidationError(
                    "main_code is required"
                )

            try:

                main_service = MainService.objects.get(
                    code=main_code
                )

            except MainService.DoesNotExist:

                raise ValidationError(
                    f"Main service not found: {main_code}"
                )

            slug = slugify(title_en)

            obj, is_created = Service.objects.update_or_create(
                slug=slug,
                defaults={
                    "main_service": main_service,
                    "title_en": title_en,
                    "title_ar": title_ar,
                    "short_description_en": short_description_en,
                    "short_description_ar": short_description_ar,
                    "order": order,
                    "is_active": True,
                }
            )

            if is_created:
                created += 1

        return created

    # =====================================================
    # SECTIONS
    # =====================================================

    def import_sections(self):

        rows = self.get_rows("sections")

        created = 0

        for index, row in enumerate(rows, start=1):

            service_slug = str(
                row.get("service_slug", "")
            ).strip()

            title_en = str(
                row.get("title_en", "")
            ).strip()

            title_ar = str(
                row.get("title_ar", "")
            ).strip()

            subtitle_en = str(
                row.get("subtitle_en", "")
            ).strip()

            subtitle_ar = str(
                row.get("subtitle_ar", "")
            ).strip()

            content_en = str(
                row.get("content_en", "")
            ).strip()

            content_ar = str(
                row.get("content_ar", "")
            ).strip()

            order = row.get("order") or index

            try:

                service = Service.objects.get(
                    slug=service_slug
                )

            except Service.DoesNotExist:

                raise ValidationError(
                    f"Service not found: {service_slug}"
                )

            obj, is_created = ServiceSection.objects.update_or_create(
                service=service,
                title_en=title_en,
                defaults={
                    "title_ar": title_ar,
                    "subtitle_en": subtitle_en,
                    "subtitle_ar": subtitle_ar,
                    "content_en": content_en,
                    "content_ar": content_ar,
                    "order": order,
                    "is_active": True,
                }
            )

            if is_created:
                created += 1

        return created

    # =====================================================
    # EXECUTE
    # =====================================================

    @transaction.atomic
    def execute(self):

        self.load()

        main_services = self.import_main_services()

        services = self.import_services()

        sections = self.import_sections()

        return {
            "success": True,
            "main_services_created": main_services,
            "services_created": services,
            "sections_created": sections,
        }